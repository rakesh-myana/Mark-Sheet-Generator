from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import os

# Setup ChromeDriver
service = Service(executable_path="chromedriver.exe")
driver = webdriver.Chrome(service=service)

# Roll numbers
start = 1239
end = 1239
roll_numbers = [f"21JJ1A{i}" for i in range(start, end + 1)]

# Check if file exists; if not, create it
file_name = "Marks_sheet2.xlsx"
if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "firstSheet"
    # Add headers
    headers = ["S.No.", "Roll Number", "Name", "1-1", "1-2", "2-1", "2-2", "3-1", "3-2", "4-1", "CGPA"]
    ws.append(headers)
    wb.save(file_name)

# Load workbook
wb = load_workbook(file_name)
ws = wb["firstSheet"]

# Get current last row index
index = ws.max_row

try:
    for roll in roll_numbers:
        driver.get("https://jntuhresults.vercel.app/academicresult")

        input_element = driver.find_element(By.XPATH, "//input[contains(@placeholder,'Enter your hallticket no')]")
        input_element.clear()
        input_element.send_keys(roll)

        link = driver.find_element(By.XPATH, "//button[text()='Result']")
        link.click()

        cgpa = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.dark\\:border-white.w-\\[25\\%\\]'))
        )

        name = driver.find_element(By.XPATH, "//table//tr[2]/th[1]")
        name = name.text.split('\n')[1]

        results = driver.find_elements(By.XPATH, '//*[@class="dark:border-white w-[25%]"]')
        r_11 = results[0].text
        r_12 = results[1].text
        r_21 = results[2].text
        r_22 = results[3].text
        r_31 = results[4].text
        r_32 = results[5].text
        r_41 = results[6].text
        r_42 = results[7].text
        cgpa = results[8].text

        
        ws.append([index, roll, name,r_11, r_12, r_21, r_22,r_31, r_32, r_41,cgpa])
        index += 1

except Exception as e:
    print("Error:", e)

finally:
    wb.save(file_name)
    driver.quit()
